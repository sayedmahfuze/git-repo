# imports 

from openpyxl import Workbook, load_workbook 
import settings 

class CreateSpreadsheet:
    def __init__(self):
        self.db_path = settings.db_path
        self.spFile = settings.spreadsheet_name 
        self.path = self.db_path + self.spFile

    def create(self):
        workbook = Workbook()
        self.ws = workbook.active 
        self.ws.title = "My_Store"

        # Creating New Worksheets
        workbook.create_sheet("products")
        workbook.create_sheet("stocks")
        workbook.create_sheet("sales")
        workbook.create_sheet("history")
        workbook.create_sheet("bill")
        workbook.create_sheet("purchase")
        print(workbook.sheetnames)

        #Creating worksheet objects
        self.ws_pd = workbook["products"]
        self.ws_st = workbook["stocks"]
        self.ws_sa = workbook["sales"]
        self.ws_ht = workbook["history"]
        self.ws_bl = workbook["bill"]
        self.ws_pc = workbook["purchase"]

        # adding columns to the work sheet
        self.ws_pd["A1"] = "product_id"
        self.ws_pd["B1"] = "product_name"
        self.ws_pd["C1"] = "product_type"
        self.ws_pd["D1"] = "product_brand"
        self.ws_pd["E1"] = "measure_unit"
        self.ws_pd["F1"] = "unit_price"
        self.ws_pd["G1"] = "product_mrp"
        self.ws_pd["H1"] = "tax_catagory"

        self.ws_st["A1"] = "product_id"
        self.ws_st["B1"] = "product_name"
        self.ws_st["C1"] = "product_type"
        self.ws_st["D1"] = "product_brand"
        self.ws_st["E1"] = "measure_unit"
        self.ws_st["F1"] = "unit_price"
        self.ws_st["G1"] = "product_mrp"
        self.ws_st["H1"] = "purchase_date"
        self.ws_st["I1"] = "exp_date"
        self.ws_st["J1"] = "purchase_quantity"
        self.ws_st["K1"] = "current_quantity"

        self.ws_bl["A1"] = "product_code"
        self.ws_bl["B1"] = "product_name"        
        self.ws_bl["C1"] = "product_brand"
        self.ws_bl["D1"] = "sale_date"
        self.ws_bl["E1"] = "sale_quantity"
        self.ws_bl["F1"] = "mrp"        
        self.ws_bl["G1"] = "discount"
        self.ws_bl["H1"] = "tax"
        self.ws_bl["I1"] = "bill_id"

        self.ws_sa["A1"] = "product_code"
        self.ws_sa["B1"] = "product_name"        
        self.ws_sa["C1"] = "product_brand"
        self.ws_sa["D1"] = "sale_date"
        self.ws_sa["E1"] = "sale_quantity"
        self.ws_sa["F1"] = "mrp"        
        self.ws_sa["G1"] = "discount"
        self.ws_sa["H1"] = "tax"
        self.ws_sa["I1"] = "sale_id"
        self.ws_sa["J1"] = "bill_id"


        self.ws_pc["A1"] = "product_code"
        self.ws_pc["B1"] = "product_name"        
        self.ws_pc["C1"] = "product_brand"
        self.ws_pc["D1"] = "purchase_date"
        self.ws_pc["E1"] = "purchase_quantity"
        self.ws_pc["F1"] = "mrp"        
        self.ws_pc["G1"] = "discount"

        






        workbook.save(self.path)



if __name__ == "__main__":
    cs = CreateSpreadsheet()
    cs.create()

