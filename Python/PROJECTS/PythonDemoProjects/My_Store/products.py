from openpyxl import load_workbook
from columns import pd_columns, pd_columns_ndx
from column_class import Product
from settings import db_path, spreadsheet_name


class Products:
    """
    Insert New records to the Product or
    Update the Existing Record
    Deleting the existing Record
    fetch the desired record
    """
    def __init__(self):
        
        self.filepath = db_path + spreadsheet_name

        self.workbook = load_workbook(self.filepath)

        self.work_sheet = self.workbook['products']
 

    def get_headers(self):
        for row in self.work_sheet.iter_rows(min_row=1, values_only=True):
            return row
        
    def get_column(self, column_name):
        for col in self.work_sheet.iter_cols(min_row=1, values_only=True):
            return col

    def get_row(self, row_number):
        for row in self.work_sheet.iter_rows(min_row=1, values_only=True):
            return row 

    def get_all_data(self, sheet_name):
        pass 
    #@staticmethod
    def get_value(self):
        self.final_values = []        
        all_columns = self.get_headers()        
        for i in all_columns:
            value = input(f"Please enter value for: {i}: ")
            self.final_values.append(value)            
        return self.final_values
                


    def get_max_row(self):
        max_row_rount = 1
        temp = 0
        for v in pd_columns_ndx.values():
            for x in self.work_sheet[v]:
                temp += 1 
            if temp > max_row_rount :
                max_row_rount = temp
                temp = 0
            else:
                temp = 0
            
        return max_row_rount
        

    def insert_to_produts(self):
        row_to_insert = self.get_max_row()
        cells = []
        for i in pd_columns_ndx.values():
            cells.append(i + str(row_to_insert+1))
        print(cells)
        values = []
        values = self.get_value()
        for i in range(0,len(cells)):
            self.work_sheet[cells[i]] = values[i]
        self.workbook.save(filename=self.filepath)
        

 
if __name__ == '__main__':
    p = Products()
    #p.get_value()
    print("Do you want to Insert New Product (Y/N)")
    flag = input("> ")
    if flag == 'Y' or flag == 'y':
        while True:
            p.insert_to_produts()
            print("Do you want inset more (Y/N)")
            uinput = input("> ")
            if uinput == 'N' or uinput == 'n':
                break 

    else:
        print("Thank you.")

