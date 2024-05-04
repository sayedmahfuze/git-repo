# Print All Columns Name

import openpyxl
from datetime import datetime

xlFile = "D:\\RoadMap\\Python\\PythonDemoProjects\\billings\\items.xlsx"

# workbook object

wb_obj = openpyxl.load_workbook(xlFile)

# active sheet object

sheet_obj = wb_obj.active

max_col = sheet_obj.max_column

# loop will print all column name

for i in range (1, max_col +1):
    cell_obj = sheet_obj.cell(row=1, column=i)
    print(cell_obj.value)



# Active sheet Object


def get_items(item_type):
    item_list = []
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row

    for i in range(2, max_row+1):
      cell_obj = sheet_obj.cell(row=i, column=2)
      #print(cell_obj.value)
      if cell_obj.value == item_type:
        item_list.append(sheet_obj.cell(row=i, column=3).value)
    return item_list 

def get_unit_price(path, cell_id):
    workbook = openpyxl.load_workbook(filename=path)
    sheet = workbook.active    
    cell_pointer = 'F'+str(cell_id)
    cell = sheet[cell_pointer]
    #print(f'The variable "cell" is {cell.value}')
    return cell.value
   
def get_selected_item_cell(item):
   sheet_obj = wb_obj.active
   max_row = sheet_obj.max_row
   for i in range(2, max_row+1):
      #print(i)
      cell_obj = sheet_obj.cell(row=i, column=3)
      #print(cell_obj.value)
      indx = i
      if cell_obj.value.strip() == item:
         return indx
      
         



def my_bill():
   print("Home Retail")
   print("Please enter customer name: ")
   customer_name = input('> ')
   now = datetime.now()
   #print(now)
   print("Please choose items: ")
   print("---------------------")
   print("Enter Item Type: ")
   item_type = input("> ")
   while item_type:
      items = (get_items(item_type))
      if items:         
         print(items)
         print("Choose Items:")
         item = input("> ")
         quentity = input("Queantity: ")
         cell = get_selected_item_cell(item)
         unit_price = get_unit_price(xlFile, cell)
         print("Do you want to more Items(Y/N): ")
         flag = input("> ")
         if flag =='N':
            print("Bill")
            print("--------------------------------------")
            print(f"Name: {customer_name}      Purchase_date: {now}")
            print(f"Item: {item}   quantity: {quentity}    unit price: {unit_price}    total: {int(unit_price) * int(quentity)}")
            break 
         else:
            print("Enter Item Type: ")
            item_type = input("> ")
      else:
         print("Entered Item is not available!")
         print("Please Enter New Item Type: ")
         item_type = input("> ")
         
      

      
      

my_bill()
#item = 'Matta'
#print(get_selected_item_cell(item))