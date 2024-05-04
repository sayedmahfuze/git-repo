# Setting the font styles of the cells 

# To customize font styles in cells, important, import the Font() Function from 
# the openpyxl.styles module 

# Program to set the font of the text.

import openpyxl 
from openpyxl.styles import Font 


wb = openpyxl.Workbook()

sheet = wb.active 

sheet.cell(row=1, column=1).value = "Sayed Mahfuze"
# set the size of the cell to 24 
sheet.cell(row=1, column=1).font = Font(size=24)

sheet.cell(row=2, column=2).value = "Sabiha Naaz"
# set the font style to italic 
sheet.cell(row=2, column=2).font = Font(size=20, italic = True)

sheet.cell(row=3, column=3).value = "Khabib Nurmagomedov"
# set the font style to bold 
sheet.cell(row=3, column=3).font = Font(size=24, bold=True)

sheet.cell(row=4, column=4).value = "Hiba Nurmagomedov"
# set the font name to 'Times New Roman'
sheet.cell(row=4, column=4).font = Font(size = 20, name='Times New Roman')

wb.save('C:\\Users\\Sayed\\Desktop\\RoadMap\\Python\\23_OpenPyExcel\\font_text.xlsx')
