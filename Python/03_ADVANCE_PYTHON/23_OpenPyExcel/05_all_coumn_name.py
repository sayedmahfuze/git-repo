# Print All Columns Name

import openpyxl

xlFile = "C:\\Users\\Sayed\\Desktop\\RoadMap\\Python\\23_OpenPyExcel\\Country Capital.xlsx"

# workbook object

wb_obj = openpyxl.load_workbook(xlFile)

# active sheet object

sheet_obj = wb_obj.active

max_col = sheet_obj.max_column

# loop will print all column name

for i in range (1, max_col +1):
    cell_obj = sheet_obj.cell(row=1, column=i)
    print(cell_obj.value)
