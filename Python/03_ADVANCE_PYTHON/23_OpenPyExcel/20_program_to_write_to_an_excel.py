# Program to write to an Excel sheet 

from openpyxl import Workbook

xlFile = "C:\\Users\\Sayed\\Desktop\\RoadMap\\Python\\23_OpenPyExcel\\NewExcelWrite.xlsx"

wb = Workbook() 

sheet = wb.active 
sheet.cell(row=1, column=1).value = "ANKIT"

c2 = sheet.cell(row =1, column=2).value = "RAI"


# Once have a worksheet object, once can 
# access a cell object by its name also.
# A2 menas column = 1 & row = 2.
sheet['A2'] = "RAHUL"

# B2 nmeans column = 1 and row = 2 

sheet['B2'] = "RAI"

# anytime we modify the Workbook object
# ot its sheets and cells, the spreadsheet
# file will not be saved until we call
# the save() workbook method 

wb.save(xlFile)  


