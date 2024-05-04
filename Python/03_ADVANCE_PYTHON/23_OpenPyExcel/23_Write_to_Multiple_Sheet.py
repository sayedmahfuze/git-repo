#!/usr/bin/python

from openpyxl import Workbook
import datetime as dt 




book = Workbook()
sheet = book.active

sheet['A1'] = 'ITEM_CODE'
sheet['B1'] = "ITEM_NAME"
sheet['C1'] = "BRAND"
sheet['D1'] = "UNIT_PRICE"
sheet['E1'] = "TOTAL_UNIT"
sheet['F1'] = "PURCHASE_DATE"
sheet['G1'] = "AVAILABLE_STOCK"

mRow = sheet.max_row
mColumn = 1
sDate = dt.datetime.now()

while True:
	item = input("Item name: ")
	brand = input("Brand: ")
	unit_price = int(input("Unit Price: "))
	total_unit = int(input("Total Unit: "))
	for i in range(1, 2):
		mRow +=1
		sheet.cell(row=mRow, column=1).value = item[0:2].upper() + brand[0:2].upper() + str(mRow)
		sheet.cell(row=mRow, column=2).value = item.title() 
		sheet.cell(row=mRow, column=3).value = brand.title() 
		sheet.cell(row=mRow, column=4).value = format(unit_price,'.2f')
		sheet.cell(row=mRow, column=5).value = total_unit
		sheet.cell(row=mRow, column=6).value = sDate
		sheet.cell(row=mRow, column=7).value = total_unit

	print("Do you want to insert more: (Y/N)")
	uInput = input("> ")
	if uInput == 'Y' or uInput =='n':
		break


book.save('write_cells.xlsx')  
