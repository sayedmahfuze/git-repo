# Program to set the dimensions of the Cells 

import openpyxl 

# Call a workbook() function of openpyxl
# to create a new blank workbook object 

wb = openpyxl.Workbook()

# Get workbook active sheet 
# From the active attribute.

sheet = wb.active 

# Write to the specified cell 

sheet.cell(row=1, column=2).value = "Hello"

sheet.cell(row=2, column=2).value = "Everyone"

# Set the height of the row 

sheet.row_dimensions[1].height=70

# set the width of the column 

sheet.column_dimensions['B'].width = 20 


# save the file 

wb.save("C:\\Users\\Sayed\\Desktop\\RoadMap\\Python\\23_OpenPyExcel\\dimension.xlsx")

