from json import load
from openpyxl import load_workbook 
import data_file_config as dfc

class Spreadsheet:
    def __init__(self, work_sheet):
        self.file = dfc.file_os_path         
        self.work_book = load_workbook(filename=self.file)
        self.work_sheet = work_sheet
        self.alphabets = dfc.alphabets
        

    def check_sheet_name(self, sheet_name):
        """
        available_sheets = self.work_book.sheetnames
        print(available_sheets)
        if sheet_name in available_sheets:
            self.work_sheet = self.work_book[sheet_name]
        else:
            print(f"Incorrect Sheet name {sheet_name}")
        """
        try:
            self.work_sheet = self.work_book[sheet_name]
            print("connection successfull")
            return self.work_sheet
        except:
            while True:
                print("Please insert  a Valid sheet name: ")
                self.sheet_name = input("> ")
                try:
                    self.work_sheet = self.work_book[self.sheet_name]
                    print("connection successfull")
                    break
                except:
                    continue
            return self.work_sheet



    def get_count(self):
        ws = self.check_sheet_name(self.work_sheet)
        all_rows = []
        mx_row = 0
        for i in ws.iter_rows(min_row = 1, max_row = 1, values_only=True):
            for column_name in i:
                all_rows.append(column_name)
        map_columns = list(zip(self.alphabets, all_rows ))
        for i in map_columns:
            all_rows = ws[i[0]]
            if len(all_rows) > mx_row:
                 mx_row = len(all_rows)
        return ws, mx_row, map_columns

    def insert_new_record(self):
        #ws = self.check_sheet_name(self.work_sheet)
        ws, max_row, columns = self.get_count()
        max_prod_id = ws[columns[0][0]+str(max_row)].value
        print(max_prod_id)
        print("Do you want to insert new records (Y/N)")
        flag = input("> ")
        if flag == 'Y' or flag =='y':
            while True:
                value_list = []
                max_row += 1
                print("Please Give the Bellow Details")
                # PRODUCT_ID	PRODUCT_NAME	PRODUCT_CATAGORY	SUB_CATAGORY	BRAND	MEASURE_UNIT	TAX_SLAB	EXPIRABLE
                for i in columns[:]:
                    print(i)
                    inValue = input(f"{i[1]}: ")
                    value_list.append(inValue)
                print("Do you want to insert more: (Y/N)")
                inFlag = input("> ")
                if inFlag == 'Y' or inFlag =='y':
                    try:
                        for i in range(0, len(columns)):
                            ws[columns[i][0]+ str(max_row)] = value_list[i]
                            print(f"Column: {columns[i][0]+ str(max_row)} value {value_list[i]}")
                        self.work_book.save(filename= self.file)
                        continue
                    except:
                        print("Something Went Wrong, Please Re-Try...")
                        continue
                    
                     
                else:
                    try:
                        for i in range(0, len(columns)):
                            ws[columns[i][0]+ str(max_row)] = value_list[i]
                            print(f"Column: {columns[i][0]+ str(max_row)} value {value_list[i]}")
                            self.work_book.save(filename= self.file)

                        break
                    except:
                        print("Something Went Wrong, unable to save the Data.")
                        break
                
                
                
        else:
            exit()
           
            
        
            


if __name__ == "__main__":
    s = Spreadsheet('Products')   
    s.insert_new_record()
    