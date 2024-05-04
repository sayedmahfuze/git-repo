# Merging the cells:
"""
A rectangular area of cells can be merged into a single cell with the 
merger_cells() sheet method. The argument to merge_cells() is a single 
string of the top-let and bottom-right cells of the rectangular area to be 
merged.

"""

# Program to merge the cells.
 
import openpyxl 

wb = openpyxl.Workbook()

sheet = wb.active 

sheet.merge_cells('A2:D4')

sheet.cell(row = 2, column = 1).value='Twelve cells join together.'

# merge cell C6 and D6

sheet.merge_cells('C6:D6')
sheet.cell(row = 6, column = 6).value = 'Two merge cells.'

wb.save("C:\\Users\\Sayed\\Desktop\\RoadMap\\Python\\23_OpenPyExcel\\merge.xlsx")

