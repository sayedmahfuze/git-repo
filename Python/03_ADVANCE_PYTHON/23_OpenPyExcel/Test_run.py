from openpyxl import load_workbook 

file_os_path = "D:\\Users\\MRaheman\\Desktop\\Practice\\general_products.xlsx"
alphabets =  [chr(value).upper() for value in range(97, 123)]

wb = load_workbook(filename = file_os_path)

ws = wb['Products']

print(ws['A1'].value)