# Unmerging the cells:
# To unmerge cells, call the unmerge_cells() sheet method.

import openpyxl 

xlFile = "C:\\Users\\Sayed\\Desktop\\RoadMap\\Python\\23_OpenPyExcel\\merge.xlsx"
wb = openpyxl.load_workbook(xlFile)

sheet = wb.active 

# unmerge the cells 

sheet.unmerge_cells('A2:D4')

sheet.unmerge_cells('C6:D6')

wb.save(xlFile)

