# Print A Perticular Row Value

import openpyxl

# File path

xlFile = "C:\\Users\\Sayed\\Desktop\\RoadMap\\Python\\23_OpenPyExcel\\Country Capital.xlsx"


# work book Object

wb_obj = openpyxl.load_workbook(xlFile)

# active sheet object

sheet_obj = wb_obj.active

max_col = sheet_obj.max_column

for i in range(1, max_col+1):
  cell_obj = sheet_obj.cell(row=2, column=i)
  print(cell_obj.value, end=" ")
