# Print First column Value

import openpyxl

# File path

xlFile = "C:\\Users\\Sayed\\Desktop\\RoadMap\\Python\\23_OpenPyExcel\\Country Capital.xlsx"

# workbook object

wb_obj = openpyxl.load_workbook(xlFile)

# Active sheet Object

sheet_obj = wb_obj.active

max_row = sheet_obj.max_row

for i in range(1, max_row+1):
  cell_obj = sheet_obj.cell(row=i, column=1)
  print(cell_obj.value)
