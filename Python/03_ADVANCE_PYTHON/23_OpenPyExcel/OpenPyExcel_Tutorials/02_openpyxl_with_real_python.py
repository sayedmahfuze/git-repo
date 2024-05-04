# Reading Execl Spreadsheets with openpyxl:

# A Simple Approach to Reading an Excel Spreadsheet 

from openpyxl import load_workbook 

path = "D:\\RoadMap\\Python\\23_OpenPyExcel\\OpenPyExcel_Tutorials\\"

wb = load_workbook(filename = path+"Ex01.xlsx")

print(wb.sheetnames)

wb_p = wb.active 

print(wb_p)

print(wb_p.title)

# To Retrive Data from a  Cell 

print(wb_p["A1"].value)

# We can alos use the method .cell() to retrive a cell using index notation.
# Note: always add .value to get the actual value and not a Cell object.

cell = wb_p.cell(row=1, column=2)
print(cell)
print(cell.value)


cell_value = wb_p.cell(row=1, column=1).value
print(f"cell value: {cell_value}")


# Note: Even though in Python we are using used to a zero-indexed notaion, with the spreadsheet we will always use one-indexed notaion 
#       where the first row or column always has index 1.

# Additional Reading Options 

# There are few arguments we can pass to load_workbook() that change the way a spreadsheet is loaded.
# The most important once are the following two.
# 1. read_only : loads a spreadsheet in read-only mode allowing us to open vary large excel files.
# 2. data_only : ignores loading formulas and instead loads only the resulting values.

