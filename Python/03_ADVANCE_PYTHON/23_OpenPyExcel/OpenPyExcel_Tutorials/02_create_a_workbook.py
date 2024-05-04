# Create a workbook 
# There is no need to create a file on the filesystem to get started with openpyxl.
# Just import the Workbook class and start work.

from openpyxl import Workbook 

wb = Workbook()

# A workbook is always created with at least one worksheet.
# We can get it by using the Workbook.active property.

ws = wb.active 

# Note: This is set to 0 by default. Unless we modify its value, we will
# always get the first worksheet by using this method.

# We can create new worksheets using the Workbook.create_sheet() method:

ws1 = wb.create_sheet("product")  # insert at the end (default)

ws2 = wb.create_sheet("sales", 0) # insert at first position 

ws3 = wb.create_sheet("stock", -1) # insert at the penultimate position 

# Sheets are given a name automatically when they are created.
# They are numbered in sequence (Sheet, Sheet1, Sheet2...).
# We can change this name at any time with the Worksheet.title property.

ws.title = "New Title"

# Once we gave a worksheet a name, we can get it as a key of the workbook 

ws4 = wb["New Title"]

# We can review the names of all worksheets of the workbook with the Workbook.sheetname attribute 

print(wb.sheetnames)

# We can loop through worksheets

for sheet in wb:
    print(sheet.title)

# we can create a copies of worksheets within a single workbook:
# Workbook.copy_worksheet() method

source = wb.active 
target = wb.copy_worksheet(source)

# Note: 
"""
Only cells (including vaues, styles, hyperlinks and comments) and certain worksheet attributes
(including dimensions, format and properties) are copied. All other workbook / worksheet
attributes are not copied e.g. Images, Charts.

We also cannot copy worksheets between workbooks.
We cannot copy a worksheet if the workbook is open in read-only or write-only mode.
"""

wb.save("D:\\RoadMap\\Python\\23_OpenPyExcel\\OpenPyExcel_Tutorials\\02_excel.xlsx")

