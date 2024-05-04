# Loading from a File 

# we can use the openpyxl.load_workbook() to open an existing workbook 

from openpyxl import load_workbook

wb = load_workbook(filename= "D:\\RoadMap\\Python\\23_OpenPyExcel\\OpenPyExcel_Tutorials\\04_excel.xlsx")

sheet_range = wb['unit']

print(sheet_range["B2"].value)                                              


