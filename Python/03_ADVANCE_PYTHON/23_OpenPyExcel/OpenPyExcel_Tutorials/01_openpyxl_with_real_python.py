# Learning Some Basic Excel Terminology.

"""
Spreadsheet or Workbook:    A Spreadsheet is the main file we are creating or working with.
Worksheet or Sheet:         A Sheet is used to split different kinds of content within the same spreadsheet. A Spreadsheet can have one or more Sheets.
Column:                     A Column is a vertical line, and it's represented by an uppercase letter: A.
Row:                        A Row is a horizontal line, and it's represented by a number: 1.
Cell:                       A Cell is a combination if Column and Row, represented by both an uppercase letter and a number: A1.

"""

from openpyxl import Workbook 

path = "D:\\RoadMap\\Python\\23_OpenPyExcel\\OpenPyExcel_Tutorials\\"

wb = Workbook()

ws = wb.active

ws.title = "Products"

ws["A1"] = "Prd_ID"
ws["B1"] = "Prd_Name"

wb.save(filename=path+"Ex01.xlsx")


