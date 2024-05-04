# Playing with Data 

from openpyxl import Workbook 

# Workbook save path 
path = "D:\\RoadMap\\Python\\23_OpenPyExcel\\OpenPyExcel_Tutorials\\"

wb = Workbook()

sheet = wb.active  # 

sheet.title = "products" # Renaming the initial active sheet

sheet1 = wb.create_sheet("stocks") # adding New work sheet 

for i in wb.sheetnames:
    print(i) # print all the sheets available 


ws_p = wb["products"] # creating a object of a particular "products" sheet 
ws_s = wb["stocks"] # creating a object of a particular "stocks" sheet 

# Accessing one cell from Product sheet 

c = ws_p["A4"] # This will return the cell at A4, or create one if it does not exists yet.

# values can be directly assigned.

ws_p["A1"] = "product_id"
ws_p["B1"] = "Product_name"
ws_p["C1"] = "product_price"

# there is also the Worksheet.cell() method 
# This provides accesss to cells using row and column notation:
ws_p.cell(row=1, column=4, value="unit")

# adding headers to the stock sheet 

ws_s.cell(row=1, column=1, value="stock_id")
ws_s.cell(row=1, column=2, value="stock_name")
ws_s.cell(row=1, column=3, value="stock_category")
ws_s.cell(row=1, column=4, value="unit")
ws_s.cell(row=1, column=5, value="unit_price")



wb.save(path+"03_excel.xlsx")


# Note: When a worksheet is created in memory, it contains no cells, They are created when first accessed.
