# Managing Sheets 

from openpyxl import Workbook 

file_path = "D:\\RoadMap\\Python\\23_OpenPyExcel\\OpenPyExcel_Tutorials\\"
file_name = file_path + '10_sample.xlsx'

# function to view the all cell values of a sheet
def print_rows(sheet_name):
    for row in sheet_name.iter_rows(values_only=True):
        print(row)

wb = Workbook() # Create a new empty workbook

wb.create_sheet("Products", 0) # create a new sheet named as Sample
wb.create_sheet("Company Sales", 1)

ws_p = wb["Products"]
ws_c = wb["Company Sales"]

ws_p["A1"] = "Product_ID" # add data to a specific cell
ws_p["B1"] = "Product_NAME" # add data to a specific cell
ws_p["C1"] = "Product_TYPE" # add data to a specific cell
ws_p["D1"] = "Product_BRAND" # add data to a specific cell

ws_c["A1"] = "Sales_ID" # add data to a specific cell
ws_c["B1"] = "Customer_NAME" # add data to a specific cell
ws_c["C1"] = "Item_QUANTITY" # add data to a specific cell
ws_c["D1"] = "Item_NAME" # add data to a specific cell

wb.save(filename = file_name) # Save the spreadsheet when we are done.


print(wb.sheetnames)

# We can select a sheet using its title 

product_sheet = wb["Products"]
sales_sheet = wb["Company Sales"]
new_sheet = wb["Sheet"]

new_sheet.title = "New Products"

print(wb.sheetnames)


# If we want to create or delete sheets, then we can alos do that with 
# .create_sheet() and .remove()

operation_sheet = wb.create_sheet("Operations")
print(wb.sheetnames)

# We can also define the position to create the sheet at 

hr_sheet = wb.create_sheet("HR", 0)
print(wb.sheetnames)


# To Remove then Just pass the sheet as an argument to the .remove()

wb.remove(operation_sheet)
print(wb.sheetnames)

wb.remove(hr_sheet)
print(wb.sheetnames)

# One other thing we can do is make duplicates of a sheet using copy_worksheet()

wb.copy_worksheet(product_sheet)
print(wb.sheetnames)