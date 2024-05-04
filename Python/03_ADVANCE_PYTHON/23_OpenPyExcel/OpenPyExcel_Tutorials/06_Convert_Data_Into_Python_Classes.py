# Convert Data Into Python Classes 
# Here we will be using the new Python Data Classes available from 
# Python 3.7., if we are using an older version of Python then we can use the 
# default Classes instead.




from openpyxl import Workbook, load_workbook

# Path of the Excel 
path = "D:\\RoadMap\\Python\\23_OpenPyExcel\\OpenPyExcel_Tutorials\\"

# Create a Sample Spreadsheet 

wb = Workbook()

ws = wb.active
ws.title = "Products"
wb.create_sheet("Stocks")
wb.create_sheet("gst")
wb.create_sheet("unit")

print(wb.sheetnames)


ws_p = wb['Products']
ws_s = wb['Stocks']
ws_g = wb["gst"]
ws_u = wb["unit"]


# Product Sheet Column Headers 
ws_p.cell(row=1, column=1, value='product_id')
ws_p.cell(row=1, column=2, value='product_name')
ws_p.cell(row=1, column=3, value='product_type')
ws_p.cell(row=1, column=4, value='brand')
ws_p.cell(row=1, column=5, value='unit_id')
ws_p.cell(row=1, column=6, value='unit_price')
ws_p.cell(row=1, column=7, value='mrp')
ws_p.cell(row=1, column=8, value='margin')
ws_p.cell(row=1, column=9, value='purchase_date')
ws_p.cell(row=1, column=10, value='expiry_date')
ws_p.cell(row=1, column=11, value='description')
ws_p.cell(row=1, column=12, value='taxable')

ws_s.cell(row=1, column=1, value="product_id")
ws_s.cell(row=1, column=2, value='product_name')
ws_s.cell(row=1, column=3, value='Quantity')
ws_s.cell(row=1, column=4, value='entry_date')
ws_s.cell(row=1, column=5, value='update_date')
ws_s.cell(row=1, column=6, value='unit_price')
ws_s.cell(row=1, column=7, value='unit_id')
ws_s.cell(row=1, column=8, value='unit_price')
ws_s.cell(row=1, column=9, value='total_value')
ws_s.cell(row=1, column=10, value='total_sale_qty')
ws_s.cell(row=1, column=11, value='expiry_date')
ws_s.cell(row=1, column=12, value='status')


ws_g.cell(row=1, column=1, value = "gst_id")
ws_g.cell(row=1, column=2, value = "start_range")
ws_g.cell(row=1, column=3, value = "end_range")
ws_g.cell(row=1, column=4, value = "percentage")

ws_u.cell(row=1, column=1, value = "unit_id")
ws_u.cell(row=1, column=2, value = "type")
ws_u.cell(row=1, column=3, value = "start_range")
ws_u.cell(row=1, column=4, value = "end_range")

# Values:
ws_p.cell(row=2, column=1, value='101')
ws_p.cell(row=2, column=2, value='Rice')
ws_p.cell(row=2, column=3, value='Basmati Rice')
ws_p.cell(row=2, column=4, value='India Gate')
ws_p.cell(row=2, column=5, value='3')
ws_p.cell(row=2, column=6, value='80')
ws_p.cell(row=2, column=7, value='130')
ws_p.cell(row=2, column=8, value='20')
ws_p.cell(row=2, column=9, value='13/08/2023')
ws_p.cell(row=2, column=10, value='12/02/2024')
ws_p.cell(row=2, column=11, value='Basmati Biryani Rice')
ws_p.cell(row=2, column=12, value='N')

ws_p.cell(row=3, column=1, value='102')
ws_p.cell(row=3, column=2, value='Rice')
ws_p.cell(row=3, column=3, value='Swarna Rice')
ws_p.cell(row=3, column=4, value='India Gate')
ws_p.cell(row=3, column=5, value='3')
ws_p.cell(row=3, column=6, value='40')
ws_p.cell(row=3, column=7, value='80')
ws_p.cell(row=3, column=8, value='25')
ws_p.cell(row=3, column=9, value='13/08/2023')
ws_p.cell(row=3, column=10, value='12/08/2024')
ws_p.cell(row=3, column=11, value='Swarna Parboiled Rice')
ws_p.cell(row=3, column=12, value='N')

ws_p.cell(row=4, column=1, value='103')
ws_p.cell(row=4, column=2, value='Pulses')
ws_p.cell(row=4, column=3, value='Masoor')
ws_p.cell(row=4, column=4, value=None)
ws_p.cell(row=4, column=5, value='3')
ws_p.cell(row=4, column=6, value='90')
ws_p.cell(row=4, column=7, value='110')
ws_p.cell(row=4, column=8, value='10')
ws_p.cell(row=4, column=9, value='13/08/2023')
ws_p.cell(row=4, column=10, value='12/02/2024')
ws_p.cell(row=4, column=11, value='Masoor Daal without Chilka')
ws_p.cell(row=4, column=12, value='N')

ws_p.cell(row=5, column=1, value='104')
ws_p.cell(row=5, column=2, value='Pulses')
ws_p.cell(row=5, column=3, value='Toor')
ws_p.cell(row=5, column=4, value="Tata")
ws_p.cell(row=5, column=5, value='3')
ws_p.cell(row=5, column=6, value='120')
ws_p.cell(row=5, column=7, value='140')
ws_p.cell(row=5, column=8, value='10')
ws_p.cell(row=5, column=9, value='13/08/2023')
ws_p.cell(row=5, column=10, value='12/02/2024')
ws_p.cell(row=5, column=11, value='Harad Daal without Chilka')
ws_p.cell(row=5, column=12, value='N')

ws_p.cell(row=6, column=1, value='105')
ws_p.cell(row=6, column=2, value='Pulses')
ws_p.cell(row=6, column=3, value='Moong Dal')
ws_p.cell(row=6, column=4, value="Tata")
ws_p.cell(row=6, column=5, value='3')
ws_p.cell(row=6, column=6, value='150')
ws_p.cell(row=6, column=7, value='229')
ws_p.cell(row=6, column=8, value='15')
ws_p.cell(row=6, column=9, value='17/08/2023')
ws_p.cell(row=6, column=10, value='12/02/2024')
ws_p.cell(row=6, column=11, value='Moong Dal')
ws_p.cell(row=6, column=12, value='N')

ws_p.cell(row=7, column=1, value='106')
ws_p.cell(row=7, column=2, value='Pulses')
ws_p.cell(row=7, column=3, value='Mealmaker')
ws_p.cell(row=7, column=4, value="Saffola")
ws_p.cell(row=7, column=5, value='1')
ws_p.cell(row=7, column=6, value='65')
ws_p.cell(row=7, column=7, value='130')
ws_p.cell(row=7, column=8, value='15')
ws_p.cell(row=7, column=9, value='17/08/2023')
ws_p.cell(row=7, column=10, value='12/02/2024')
ws_p.cell(row=7, column=11, value='Saffola Mealmaker Soya Chunks')
ws_p.cell(row=7, column=12, value='N')

ws_p.cell(row=8, column=1, value='107')
ws_p.cell(row=8, column=2, value='Pulses')
ws_p.cell(row=8, column=3, value='Kabuli Chana')
ws_p.cell(row=8, column=4, value="Organic Tattva")
ws_p.cell(row=8, column=5, value='1')
ws_p.cell(row=8, column=6, value='125')
ws_p.cell(row=8, column=7, value='175')
ws_p.cell(row=8, column=8, value='15')
ws_p.cell(row=8, column=9, value='18/08/2023')
ws_p.cell(row=8, column=10, value='12/02/2024')
ws_p.cell(row=8, column=11, value='Organic Tattva Kabuli Chana')
ws_p.cell(row=8, column=12, value='N')

ws_p.cell(row=9, column=1, value='108')
ws_p.cell(row=9, column=2, value='Pulses')
ws_p.cell(row=9, column=3, value='Urad')
ws_p.cell(row=9, column=4, value="Tenali")
ws_p.cell(row=9, column=5, value='3')
ws_p.cell(row=9, column=6, value='157')
ws_p.cell(row=9, column=7, value='180')
ws_p.cell(row=9, column=8, value='15')
ws_p.cell(row=9, column=9, value='19/09/2023')
ws_p.cell(row=9, column=10, value='12/02/2024')
ws_p.cell(row=9, column=11, value='Tenali Double Horse Urad Gota')
ws_p.cell(row=9, column=12, value='N')

ws_p.cell(row=10, column=1, value='109')
ws_p.cell(row=10, column=2, value='Pulses')
ws_p.cell(row=10, column=3, value='Green Moong')
ws_p.cell(row=10, column=4, value="24 Mantra")
ws_p.cell(row=10, column=5, value='3')
ws_p.cell(row=10, column=6, value='112')
ws_p.cell(row=10, column=7, value='125')
ws_p.cell(row=10, column=8, value='15')
ws_p.cell(row=10, column=9, value='110/010/2023')
ws_p.cell(row=10, column=10, value='12/02/2024')
ws_p.cell(row=10, column=11, value='24 Mantra Organic Green Moong Whole 500g')
ws_p.cell(row=10, column=12, value='N')

ws_p.cell(row=11, column=1, value='110')
ws_p.cell(row=11, column=2, value='Pulses')
ws_p.cell(row=11, column=3, value='Peanut')
ws_p.cell(row=11, column=4, value="vsr")
ws_p.cell(row=11, column=5, value='3')
ws_p.cell(row=11, column=6, value='208')
ws_p.cell(row=11, column=7, value='238')
ws_p.cell(row=11, column=8, value='15')
ws_p.cell(row=11, column=9, value='111/011/2023')
ws_p.cell(row=11, column=10, value='12/02/2024')
ws_p.cell(row=11, column=11, value='VSR Raw Peanut 1Kg')
ws_p.cell(row=11, column=12, value='N')



wb.save(filename=path+"05_sample.xlsx")

# so here we can extract two significant elelments from the data
# . Product
# . Price 

# A product has 
# . ID
# . Name
# . product_type
# . Brand
# . description 

# A Price has
# . Prd Name 
# . Price
# . Exp Date 
# . taxable 


# So a Straightforward implementation of these two class could be written in a separate file classes.py:

import datetime 
from dataclasses import dataclass 

@dataclass 
class Product:
    id: str 
    name: str 
    type : str
    brand: str 
    description: str 

@dataclass 
class Price:
    name: str     
    price: int 
    exp_date: datetime.datetime
    taxable: str 

# After defining our data classes, we need to convert the data from the 
# spreadsheet into these new structures.

for values in ws_p.iter_rows(min_row=1, max_row=1, values_only=True):
    print(values)

# or an alternatice 
for cell in ws_p[1]:
    print(cell.value)


# Create a file mapping.py where we have a list of all the fields names and their column location
# (zero-index) on the spreadsheet.
"""
# Products fields:
product_id =  0
product_name = 1
product_type = 2
brand = 3 
description = 10

# Price Fields 
#product_name = 2
product_price = 5
exp_date = 9
taxable = 11
"""

from datetime import datetime 
from openpyxl import load_workbook
from classes import Product, Price
from mapping import product_id, product_name, product_type, brand, \
    description, product_price, exp_date, taxable

# Using the read_only method since we are not gonna be editing the spreadsheet

workbook = load_workbook(filename=path+"05_sample.xlsx", read_only=True)
sheet = workbook["Products"]

products = []
prices = []

# using the values_only because we just want tp return the cell value
for row in sheet.iter_rows(min_row=2, values_only=True):
    product = Product( id = row[product_id], 
                       name = row[product_name],
                       type = row[product_type],
                       brand = row[brand],
                       description = row[description] )
    products.append(product)

    # We need to parse the date from the spreadsheet into a datetime format
    spread_date = row[exp_date]
    parsed_date = datetime.strptime(spread_date, "%d/%m/%Y")
    price = Price(name = row[product_type],
                   price = row[product_price],
                   exp_date = row[exp_date],
                   taxable = row[taxable])
    
    prices.append(price)

print(products[0])
print(prices[0])

