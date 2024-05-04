# Managing Rows and Columns

"""
One of the most common things we have to do when manipulating spreadsheets in adding or 
removing rows and columns. The openpyxl packages allows us to do that in a very 
straightforward way by using the methods:
. insert_rows()
. delete_rows()
. insert_cols()
. delete_cols()

Every single one of those methods can receive two arguments:

1. idx
2. amount

"""

from openpyxl import Workbook 

file_path = "D:\\RoadMap\\Python\\23_OpenPyExcel\\OpenPyExcel_Tutorials\\"
file_name = file_path + '09_sample.xlsx'

# function to view the all cell values of a sheet
def print_rows(sheet_name):
    for row in sheet_name.iter_rows(values_only=True):
        print(row)

wb = Workbook() # Create a new empty workbook

wb.create_sheet("Sample", 0) # create a new sheet named as Sample

ws = wb["Sample"]

ws["A1"] = "Hello" # add data to a specific cell
ws["B1"] = "World" # add data to a specific cell

wb.save(filename = file_name) # Save the spreadsheet when we are done.


print_rows(ws)

# Insert a column before the existing column 1 ("A")
ws.insert_cols(idx=1)

print_rows(ws)

# Insert 5 columns between column 2 ("B") and 3 ("C")
ws.insert_cols(idx=3, amount=5)
print_rows(ws)

# Deelte the created columns 
ws.delete_cols(idx=3, amount=5)
ws.delete_cols(idx=1)
print_rows(ws)


# Insert a new row in the beginning 

ws.insert_rows(idx=1)
print_rows(ws)

# Insert 3 rows in the beginning
ws.insert_rows(idx=1, amount=3)
print_rows(ws)

# Delete the first 4 rows
ws.delete_rows(idx=1, amount=4)
print_rows(ws)


"""
The only thing we need to remember is that when inserting new data(rows or columns), the insertion happens before the idx parameter.
So if we do insert_rows(1), it insert a new row before the existing first row.

It's the same for columns: when we call insert_cols(2), it inserts a new column right before the already existing second column (b).
However, when deleteting rows or columns, delete_... deletes data starting from the index passed as an argument.

"""

