# Writing Excel Spreadsheets with openpyxl 

"""
There are a lot of different things we can write to a spreadsheet,
from simple text or number values to complex formulas, charts, or 
even images.
"""

from openpyxl import Workbook 

file_path = "D:\\RoadMap\\Python\\23_OpenPyExcel\\OpenPyExcel_Tutorials\\"
file_name = file_path + '08_sample.xlsx'

# function to view the all cell values of a sheet
def print_rows(sheet_name):
    for row in sheet_name.iter_rows(values_only=True):
        print(row)

wb = Workbook() # Create a new empty workbook

wb.create_sheet("Sample", 0) # create a new sheet named as Sample

ws = wb["Sample"]

ws["A1"] = "Hello" # add data to a specific cell
ws["B1"] = "World" # add data to a specific cell

wb.save(filename = file_name) # Save the spreadsheet when we are done.

print(wb.sheetnames)

print_rows(ws)


# Addding and Updating Cell Values 

# we already learned how to add values to a spreadsheet like 

ws["C1"] = "OPENPYXL"

# There is another way we can do this, by first selecting a cell and then changing its value.

cell = ws["A1"]
print(f"Cell: {cell}")

print(f"Value: {cell.value}")

cell.value = "PYTHON"
print(f"Cell Value After Update: {cell.value}")

print_rows(ws)

# Try adding a value to row 10 

ws["B10"] = "My World"

print_rows(ws) 

