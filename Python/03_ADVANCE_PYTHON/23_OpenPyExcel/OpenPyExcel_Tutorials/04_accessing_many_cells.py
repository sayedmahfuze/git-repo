# Accessing many cells 

# Ranges of rows or columns can be accessed using slicing 

from openpyxl import  Workbook

path = "D:\\RoadMap\\Python\\23_OpenPyExcel\\OpenPyExcel_Tutorials\\"

wb = Workbook()

ws = wb.active
ws.title = "Products"
wb.create_sheet("Stocks")
wb.create_sheet("gst")
wb.create_sheet("unit")


ws_p = wb['Products']
ws_s = wb['Stocks']
ws_g = wb["gst"]
ws_u = wb["unit"]

print(wb.sheetnames)
# Product Sheet Column Headers 
ws_p.cell(row=1, column=1, value='product_id')
ws_p.cell(row=1, column=2, value='product_name')
ws_p.cell(row=1, column=3, value='product_type')
ws_p.cell(row=1, column=4, value='brand')
ws_p.cell(row=1, column=5, value='unit_id')
ws_p.cell(row=1, column=6, value='unit_price')
ws_p.cell(row=1, column=7, value='mrp')
ws_p.cell(row=1, column=8, value='margin')
ws_p.cell(row=1, column=9, value='purchase_date')
ws_p.cell(row=1, column=10, value='expiry_date')
ws_p.cell(row=1, column=11, value='description')
ws_p.cell(row=1, column=12, value='taxable')

ws_s.cell(row=1, column=1, value="product_id")
ws_s.cell(row=1, column=2, value='product_name')
ws_s.cell(row=1, column=3, value='Quantity')
ws_s.cell(row=1, column=4, value='entry_date')
ws_s.cell(row=1, column=5, value='update_date')
ws_s.cell(row=1, column=6, value='unit_price')
ws_s.cell(row=1, column=7, value='unit_id')
ws_s.cell(row=1, column=8, value='unit_price')
ws_s.cell(row=1, column=9, value='total_value')
ws_s.cell(row=1, column=10, value='total_sale_qty')
ws_s.cell(row=1, column=11, value='expiry_date')
ws_s.cell(row=1, column=12, value='status')


ws_g.cell(row=1, column=1, value = "gst_id")
ws_g.cell(row=1, column=2, value = "start_range")
ws_g.cell(row=1, column=3, value = "end_range")
ws_g.cell(row=1, column=4, value = "percentage")

ws_u.cell(row=1, column=1, value = "unit_id")
ws_u.cell(row=1, column=2, value = "type")
ws_u.cell(row=1, column=3, value = "start_range")
ws_u.cell(row=1, column=4, value = "end_range")


# accessing many cells 

# Ranges of cells can be accessed using slicing:

cell_range = ws_p['A1':'C2']

print(cell_range)


# Ranges of rows or columns can be obtained similarly:

colC = ws_p['C']
col_range = ws_p['C:D']
row10 = ws_p[10]
row_range = ws_p[5:10]
#print(f"colC: {colC}")
#print(f"col_range: {col_range}")
#print(f"row10: {row10}")
#print(f"row_range: {row_range}")

# we can also use the Worksheet.iter_rows() method 

for row in ws_p.iter_rows(min_row=1, max_col=3, max_row=5):
    for cell in row:
        print(cell) 


# Note: For performance reason the Worksheet.iter_cols() method is not available in read-only mode.

# If we need to iterate through all the rows or columns of a file,
# we can insted use the Worksheet.rows property.

# 104	Sunflower Oil	Edible Oil	Anjali	4	136	240	15	05/08/23		Anjali Sunlora sunflower Oil	Y

# Adding value to unit work sheet
ws_u['A2'] = 1
ws_u['B2'] = 'GM'
ws_u['A3'] = 2
ws_u['B3'] = 'ML'
ws_u['A4'] = 3
ws_u['B4'] = 'KG'
ws_u['A5'] = 4
ws_u['B5'] = 'LT'
ws_u['A6'] = 5
ws_u['B6'] = 'PC'

cell_rows = tuple(ws_u.rows)
print(cell_rows)

cell_columns = tuple(ws_u.columns)
print(cell_columns)

# Note: For performace reasons the Worksheet.columns property is not available in re-only mode.

# Values only 

# if we just want the values from a workshet we can use the worksheet.values property.
# This iterates over all the rows in a worksheet but returns just the cell values:

for row in ws_u.values:
    print(row)
    for value in row:
        print(value)



# Both Worksheet.iter_rows() and Worksheet.iter_cols() can take the values_only parameter to return kust the cell's value.

for row in ws_u.iter_rows(min_row=1, max_col=2, max_row=5, values_only=True):
    print(row) 

# Data Storage 

# Once we have a Cell, we can assign it a value:


# Saving to a File
# The simplest and safest way to save a workbook is by using the Workbook.save() method
# of the Workbook object.


wb.save(path + "04_excel.xlsx")

# This operation will overwrite existing files without warning.


# If required, we can specify the attribute wb.template = True, to save a workbook as template

wb.template = True 
wb.save(path + "document_template.xltx")


# Saving as a stream 
# if we want ot save the file to a stream, e.g. when using a web application
# such as Pyramid, Flask or Django then we can simply provide a NamedTemporaryFile():

"""

from tempfile import NamedTemporaryFile 
from openpyxl import Workbook 

wb = Workbook()

with NamedTemporaryFile() as tmp:
    wb.save(tmp.name)
    tmp.seek(0)
    stream = tmp.read() 

"""




