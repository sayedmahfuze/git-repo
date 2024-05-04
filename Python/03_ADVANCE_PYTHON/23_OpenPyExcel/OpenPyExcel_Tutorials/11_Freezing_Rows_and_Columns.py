# Freezing Rows and Columns 

"""
Something that we might want to do when working with big
spreadsheets is to freeze a few rows or columns, so they 
remain visible when we scroll right or down.

Freezing data allows us to keep an eye on important rows or columns, 
regardless of where we scroll in the spreadsheet.

openpyxl also has a way to accomplish this by using the worksheet freeze_panes attribute.

"""


from openpyxl import Workbook, load_workbook 

file_path = "D:\\RoadMap\\Python\\23_OpenPyExcel\\OpenPyExcel_Tutorials\\"
file_name = file_path + '11_sample.xlsx'

# function to view the all cell values of a sheet
def print_rows(sheet_name):
    for row in sheet_name.iter_rows(values_only=True):
        print(row)

wb = Workbook() # Create a new empty workbook

wb.create_sheet("Products", 0) # create a new sheet named as Sample
wb.create_sheet("Company Sales", 1)

ws_p = wb["Products"]
ws_c = wb["Company Sales"]

ws_p["A1"] = "Product_ID" # add data to a specific cell
ws_p["B1"] = "Product_NAME" # add data to a specific cell
ws_p["C1"] = "Product_TYPE" # add data to a specific cell
ws_p["D1"] = "Product_BRAND" # add data to a specific cell

ws_c["A1"] = "Sales_ID" # add data to a specific cell
ws_c["B1"] = "Customer_NAME" # add data to a specific cell
ws_c["C1"] = "Item_QUANTITY" # add data to a specific cell
ws_c["D1"] = "Item_NAME" # add data to a specific cell

wb.save(filename = file_name) # Save the spreadsheet when we are done.

wb = load_workbook(filename = file_name)

sheet = wb.active 

sheet.freeze_panes = "C1" 

wb.save(filename = file_name)

