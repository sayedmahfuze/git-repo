# Importing Data from a Spreadsheet 

# Iterating Through the Data.

# There are a few different ways we can iterate through the data depending on our needs.

# we can slice the data with a combination of columns and rows:

from openpyxl import Workbook, load_workbook

# Path of the Excel 
path = "D:\\RoadMap\\Python\\23_OpenPyExcel\\OpenPyExcel_Tutorials\\"

# Create a Sample Spreadsheet 

wb = Workbook()

ws = wb.active
ws.title = "Products"
wb.create_sheet("Stocks")
wb.create_sheet("gst")
wb.create_sheet("unit")

print(wb.sheetnames)


ws_p = wb['Products']
ws_s = wb['Stocks']
ws_g = wb["gst"]
ws_u = wb["unit"]


# Product Sheet Column Headers 
ws_p.cell(row=1, column=1, value='product_id')
ws_p.cell(row=1, column=2, value='product_name')
ws_p.cell(row=1, column=3, value='product_type')
ws_p.cell(row=1, column=4, value='brand')
ws_p.cell(row=1, column=5, value='unit_id')
ws_p.cell(row=1, column=6, value='unit_price')
ws_p.cell(row=1, column=7, value='mrp')
ws_p.cell(row=1, column=8, value='margin')
ws_p.cell(row=1, column=9, value='purchase_date')
ws_p.cell(row=1, column=10, value='expiry_date')
ws_p.cell(row=1, column=11, value='description')
ws_p.cell(row=1, column=12, value='taxable')

ws_s.cell(row=1, column=1, value="product_id")
ws_s.cell(row=1, column=2, value='product_name')
ws_s.cell(row=1, column=3, value='Quantity')
ws_s.cell(row=1, column=4, value='entry_date')
ws_s.cell(row=1, column=5, value='update_date')
ws_s.cell(row=1, column=6, value='unit_price')
ws_s.cell(row=1, column=7, value='unit_id')
ws_s.cell(row=1, column=8, value='unit_price')
ws_s.cell(row=1, column=9, value='total_value')
ws_s.cell(row=1, column=10, value='total_sale_qty')
ws_s.cell(row=1, column=11, value='expiry_date')
ws_s.cell(row=1, column=12, value='status')


ws_g.cell(row=1, column=1, value = "gst_id")
ws_g.cell(row=1, column=2, value = "start_range")
ws_g.cell(row=1, column=3, value = "end_range")
ws_g.cell(row=1, column=4, value = "percentage")

ws_u.cell(row=1, column=1, value = "unit_id")
ws_u.cell(row=1, column=2, value = "type")
ws_u.cell(row=1, column=3, value = "start_range")
ws_u.cell(row=1, column=4, value = "end_range")

# Values:
ws_p.cell(row=2, column=1, value='101')
ws_p.cell(row=2, column=2, value='Rice')
ws_p.cell(row=2, column=3, value='Basmati Rice')
ws_p.cell(row=2, column=4, value='India Gate')
ws_p.cell(row=2, column=5, value='3')
ws_p.cell(row=2, column=6, value='80')
ws_p.cell(row=2, column=7, value='130')
ws_p.cell(row=2, column=8, value='20')
ws_p.cell(row=2, column=9, value='13/08/2023')
ws_p.cell(row=2, column=10, value='12/02/2024')
ws_p.cell(row=2, column=11, value='Basmati Biryani Rice')
ws_p.cell(row=2, column=12, value='N')

ws_p.cell(row=3, column=1, value='102')
ws_p.cell(row=3, column=2, value='Rice')
ws_p.cell(row=3, column=3, value='Swarna Rice')
ws_p.cell(row=3, column=4, value='India Gate')
ws_p.cell(row=3, column=5, value='3')
ws_p.cell(row=3, column=6, value='40')
ws_p.cell(row=3, column=7, value='80')
ws_p.cell(row=3, column=8, value='25')
ws_p.cell(row=3, column=9, value='13/08/2023')
ws_p.cell(row=3, column=10, value='12/08/2024')
ws_p.cell(row=3, column=11, value='Swarna Parboiled Rice')
ws_p.cell(row=3, column=12, value='N')

ws_p.cell(row=4, column=1, value='103')
ws_p.cell(row=4, column=2, value='Pulses')
ws_p.cell(row=4, column=3, value='Masoor')
ws_p.cell(row=4, column=4, value=None)
ws_p.cell(row=4, column=5, value='3')
ws_p.cell(row=4, column=6, value='90')
ws_p.cell(row=4, column=7, value='110')
ws_p.cell(row=4, column=8, value='10')
ws_p.cell(row=4, column=9, value='13/08/2023')
ws_p.cell(row=4, column=10, value='12/02/2024')
ws_p.cell(row=4, column=11, value='Masoor Daal without Chilka')
ws_p.cell(row=4, column=12, value='N')

ws_p.cell(row=5, column=1, value='104')
ws_p.cell(row=5, column=2, value='Pulses')
ws_p.cell(row=5, column=3, value='Toor')
ws_p.cell(row=5, column=4, value="Tata")
ws_p.cell(row=5, column=5, value='3')
ws_p.cell(row=5, column=6, value='120')
ws_p.cell(row=5, column=7, value='140')
ws_p.cell(row=5, column=8, value='10')
ws_p.cell(row=5, column=9, value='13/08/2023')
ws_p.cell(row=5, column=10, value='12/02/2024')
ws_p.cell(row=5, column=11, value='Harad Daal without Chilka')
ws_p.cell(row=5, column=12, value='N')

wb.save(filename=path+"03_IDFS.xlsx")

print(ws_p["A1:L1"])

# Get all cells from column A 
print(ws_p["A"])

idx = 1
for i in ws_p["B"]:
    #print(i)
    if i.value == 'Rice':
        cell = 'C'+str(idx)
        print(ws_p[cell].value)
    idx += 1
    
# Get all cells for a range of columns 
print("\n")
print(ws_p["A:C"], end="\n")


# get alll cells from row 2
print("\n")
print(ws_p[2], end="\n")

# get all cells for a range of rows 
print("\n")
print(ws_p[3:6])

# There are also multiple ways of using normal Python generators to go through the data.
# The main methods we can use to achieve these are:

# .iter_rows()
# .item_cols()

# Both methods can receive the following arguments
# min_row 
# max_row 
# min_col
# max_col

# These arguments are used to set boundaries of the iteration:
print("\n")
for row in ws_p.iter_rows(min_row=2, max_row=5, min_col=1, max_col=7):
    print(row)

print("\n")
for col in ws_p.iter_cols(min_row=1, max_row=6, min_col=1, max_col=7):
    print(col)


# one additional argument we can pass to both methods in the boolean values_only.
# when it is set to True, the values of the cell are returned, instead of the cell objects.
print("\n")
for value in ws_p.iter_rows(min_row=2, max_row=5, min_col=1, max_col=12, values_only=True):
    print("\n")
    print(value)


# if we want ot iterate through the whole datasheet, then we can also use the attributes .row or .column directly,
# which are shortcut to using .iter_rows() and .item_cols() without any arguments.
print("\n")
for row in ws_p.rows:
    print(row)