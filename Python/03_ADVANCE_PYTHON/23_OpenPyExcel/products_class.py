from json import load
from openpyxl import Workbook, load_workbook 
import itertools 

class Products:
    def __init__(self, path ="D:\\Users\\MRaheman\\Desktop\\Practice\\", file = "Products" ):
        self.path = path
        self.file = file
        self.destination_file = self.path + self.file + '.xlsx' 
        self.alphabets = [chr(value).upper() for value in range(97, 123)]
        

    def create_new_excel(self, columns, sheet_name):
        work_book = Workbook()
        work_sheet = work_book.active 
        work_sheet.title = sheet_name
        ndx = 1

        self.column_headers = columns 
        if len(self.column_headers) > len (self.alphabets):
            #temp_alpha = self.alphabets 
            temp_alpha = []
            for i in itertools.product(self.alphabets, self.alphabets):
                temp_alpha.append(i)
            for i in temp_alpha:
                self.alphabets.append(i[0] + i[1])
        for i in range(0, len(self.column_headers)):
            work_sheet[self.alphabets[i] + str(ndx)] = self.column_headers[i]

        work_book.save(filename = self.destination_file) 

    def add_new_sheet(self, path, file_name, sheet_name):
        self.destination_file = path + file_name + '.xlsx'

        work_book = load_workbook(filename=self.destination_file)
        #print(work_book.sheetnames)
        work_book.create_sheet(sheet_name)
        work_book.save(filename = self.destination_file )
        #print(work_book.sheetnames)
    def get_max_row(self, file_name, sheet_name):
        wb = load_workbook(filename = file_name)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = []
            for i in ws.iter_rows(min_row = 1, max_row = 1, values_only=True):
                for column_name in i:
                     all_rows.append(column_name)
        map_columns = list(zip(self.alphabets, all_rows ))
        for i in map_columns:
            ws = wb[sheet_name]
            all_rows = ws[i[0]]
            print(len(all_rows))
        


    def add_values_to_all_column(self, file_name, sheet_name):
        wb = load_workbook(filename = file_name)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = []
            for i in ws.iter_rows(min_row = 1, max_row = 1, values_only=True):
                all_rows.append(i)

            
            

            
if __name__ == '__main__':
    p = Products(file = 'general_products')
    file_path ="D:\\Users\\MRaheman\\Desktop\\Practice\\"
    existing_file = "general_products"
    columns = ["PRODUCT_ID", "PRODUCT_NAME", "PRODUCT_CATAGORY", "SUB_CATAGORY", "BRAND", "MEASURE_UNIT", "TAX_SLAB", "EXPIRABLE"]
    sheet_name = "Products"
    #p.create_new_excel(columns, sheet_name)
    # new_sheet = "Srocks"
    # p.add_new_sheet(path = file_path, file_name= existing_file, sheet_name =new_sheet )
    file = file_path + existing_file + '.xlsx'
    p.get_max_row(file, sheet_name)
    

